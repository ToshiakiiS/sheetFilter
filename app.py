import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook, Workbook

EXCEL_ROW_LIMIT = 1_048_576


def get_xlsx_columns(uploaded_file):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    wb.close()
    uploaded_file.seek(0)
    return list(header) if header else []


def filter_xlsx_to_bytes(uploaded_file, selected_columns):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    if max_row and (max_row - 1) > EXCEL_ROW_LIMIT:
        wb.close()
        raise ValueError("EXCEL_ROW_LIMIT")

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    header_list = list(header) if header else []
    selected_set = set(selected_columns)
    col_indices = [i for i, name in enumerate(header_list) if name in selected_set]

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet(title=ws.title or "Sheet1")
    out_ws.append([header_list[i] for i in col_indices])

    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        out_ws.append([row[i] for i in col_indices])
        row_count += 1

    buffer = io.BytesIO()
    out_wb.save(buffer)
    buffer.seek(0)
    wb.close()
    uploaded_file.seek(0)
    return buffer, row_count, len(col_indices)

# Configure the web page
st.set_page_config(page_title="Column Selector Tool", page_icon="📊", layout="wide")

st.title("📊 Large Spreadsheet Column Filter")
st.write(
    "Upload a massive CSV or Excel file, select only the columns you want to keep, "
    "and export a clean, optimized output file."
)

# File Uploader
uploaded_file = st.file_uploader("Upload your spreadsheet (.csv, .xlsx)", type=["csv", "xlsx"])

if uploaded_file is not None:
    # Create a unique key for the file based on its name and size
    current_file_key = f"{uploaded_file.name}_{uploaded_file.size}"
    
    # OPTIMIZATION: Only parse the headers ONCE. Memorize them in session_state.
    if "file_key" not in st.session_state or st.session_state.file_key != current_file_key:
        st.session_state.file_key = current_file_key
        st.session_state.pop("output_key", None)
        st.session_state.pop("output_bytes", None)
        st.session_state.pop("output_filename", None)
        st.session_state.pop("output_mime", None)
        st.session_state.pop("output_rows", None)
        st.session_state.pop("output_cols", None)
        with st.spinner("⚡ Parsing columns for the first time..."):
            is_csv = uploaded_file.name.lower().endswith('.csv')
            if is_csv:
                header_df = pd.read_csv(uploaded_file, nrows=0)
                st.session_state.columns = header_df.columns.tolist()
            else:
                st.session_state.columns = get_xlsx_columns(uploaded_file)
            uploaded_file.seek(0)

    # Retrieve the columns safely from memory
    columns = st.session_state.columns
    
    st.write("### 🗂️ Select columns to include in the output:")
    
    # This widget will now respond INSTANTLY because it's reading a static list from memory
    selected_columns = st.multiselect(
        "Click below to add/remove columns:",
        options=columns,
        default=columns
    )
    
    if not selected_columns:
        st.warning("⚠️ Please select at least one column to generate an output.")
    else:
        st.write("### ⚙️ Export Settings")
        output_format = st.radio("Choose output format:", ["CSV", "Excel (.xlsx)"], index=1)
        output_key = (current_file_key, output_format, tuple(selected_columns))
        
        # The heavy processing only happens when this specific button is clicked
        if st.button("🚀 Process & Generate Download Link"):
            if st.session_state.get("output_key") != output_key:
                with st.spinner("Processing 800,000+ rows... Please wait."):
                    uploaded_file.seek(0)
                    is_csv = uploaded_file.name.lower().endswith('.csv')
                    buffer = io.BytesIO()

                    if output_format == "CSV":
                        if is_csv:
                            final_df = pd.read_csv(uploaded_file, usecols=selected_columns)
                        else:
                            final_df = pd.read_excel(uploaded_file, usecols=selected_columns)

                        final_df.to_csv(buffer, index=False, encoding='utf-8')
                        row_count, col_count = final_df.shape
                        mime_type = "text/csv"
                        out_filename = f"filtered_{uploaded_file.name.split('.')[0]}.csv"
                    else:
                        if is_csv:
                            final_df = pd.read_csv(uploaded_file, usecols=selected_columns)
                            if final_df.shape[0] > EXCEL_ROW_LIMIT:
                                st.error("❌ Excel files have a physical limit of 1,048,576 rows. Please use CSV instead.")
                                st.stop()

                            final_df.to_excel(buffer, index=False, engine="openpyxl")
                            row_count, col_count = final_df.shape
                        else:
                            try:
                                buffer, row_count, col_count = filter_xlsx_to_bytes(uploaded_file, selected_columns)
                            except ValueError as exc:
                                if str(exc) == "EXCEL_ROW_LIMIT":
                                    st.error("❌ Excel files have a physical limit of 1,048,576 rows. Please use CSV instead.")
                                    st.stop()
                                raise

                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        out_filename = f"filtered_{uploaded_file.name.split('.')[0]}.xlsx"

                    st.session_state.output_key = output_key
                    st.session_state.output_bytes = buffer.getvalue()
                    st.session_state.output_filename = out_filename
                    st.session_state.output_mime = mime_type
                    st.session_state.output_rows = row_count
                    st.session_state.output_cols = col_count

        if (
            st.session_state.get("output_key") == output_key
            and st.session_state.get("output_bytes")
        ):
            st.success(
                f"Done! Formatted dataset has {st.session_state.output_rows:,} rows and "
                f"{st.session_state.output_cols} columns."
            )
            st.download_button(
                label=f"📥 Download Filtered {output_format}",
                data=st.session_state.output_bytes,
                file_name=st.session_state.output_filename,
                mime=st.session_state.output_mime
            )
